import io
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_monthly_excel(user: dict, expenses: list, summary: dict, month: int, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]} {year}"

    # Colors
    primary_fill = PatternFill("solid", fgColor="1E3A5F")
    accent_fill = PatternFill("solid", fgColor="2E86AB")
    light_fill = PatternFill("solid", fgColor="EBF4F6")
    alt_fill = PatternFill("solid", fgColor="F0F7FA")

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    title_font = Font(bold=True, color="1E3A5F", size=14, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    total_font = Font(bold=True, size=11, name="Calibri")

    thin = Side(border_style="thin", color="C9DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Smart Expense Tracker — {calendar.month_name[month]} {year}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Report for: {user.get('full_name', '')}  |  Generated on: {__import__('datetime').datetime.now().strftime('%d %b %Y')}"
    ws["A2"].font = Font(size=10, color="5A5A7A", name="Calibri")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 20

    # Summary section
    ws.row_dimensions[4].height = 22
    ws["A4"] = "SUMMARY"
    ws["A4"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    ws["A4"].fill = primary_fill
    ws["A4"].alignment = center
    ws.merge_cells("A4:F4")

    summary_rows = [
        ("Total Spent", f"₹{summary.get('total_spent', 0):,.2f}"),
        ("Total Transactions", str(summary.get('transaction_count', 0))),
        ("Daily Average", f"₹{summary.get('avg_per_day', 0):,.2f}"),
        ("Total Budget", f"₹{summary.get('total_budget', 0):,.2f}"),
        ("Budget Used", f"{summary.get('budget_used_pct', 0):.1f}%"),
    ]
    for i, (label, val) in enumerate(summary_rows, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val
        ws[f"A{i}"].font = Font(bold=True, size=10, name="Calibri")
        ws[f"B{i}"].font = Font(size=10, name="Calibri")
        ws[f"A{i}"].fill = light_fill
        ws[f"B{i}"].fill = light_fill
        ws[f"A{i}"].border = border
        ws[f"B{i}"].border = border
        ws.row_dimensions[i].height = 20

    # Expenses header
    header_row = 12
    ws.row_dimensions[header_row].height = 24
    headers = ["#", "Date", "Title", "Category", "Payment Method", "Amount (₹)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = primary_fill
        cell.alignment = center
        cell.border = border

    # Column widths
    col_widths = [5, 14, 30, 16, 18, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Expense rows
    for idx, exp in enumerate(expenses, start=1):
        row = header_row + idx
        ws.row_dimensions[row].height = 20
        fill = light_fill if idx % 2 == 0 else alt_fill
        values = [
            idx,
            str(exp.get("date", ""))[:10],
            str(exp.get("title", "")),
            str(exp.get("category", "")),
            str(exp.get("payment_method", "")),
            round(float(exp.get("amount", 0)), 2),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col in [1, 4, 5, 6] else left

    # Total row
    total_row = header_row + len(expenses) + 1
    ws.row_dimensions[total_row].height = 24
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = total_font
    ws[f"A{total_row}"].fill = accent_fill
    ws[f"A{total_row}"].alignment = center

    total_cell = ws.cell(row=total_row, column=6, value=round(summary.get('total_spent', 0), 2))
    total_cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    total_cell.fill = accent_fill
    total_cell.alignment = center

    for col in range(1, 7):
        ws.cell(row=total_row, column=col).border = border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
